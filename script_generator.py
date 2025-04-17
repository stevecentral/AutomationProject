import os
from pathlib import Path

import hjson
import csv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from pkg_resources import working_set

import action_functions
from action_functions import (login, logout, power_cycle, check_element, screenshot, check_button, click, power_toggle,
                              check_text, wait, highlight, set_tile_pattern, compare_images, compare_elements,
                              send_serial_command)
from selenium import webdriver
from selenium.webdriver.common.by import By


# Initialize the WebDriver
driver = webdriver.Chrome()

# Navigate to the specified URL
address = os.getenv('ADDRESS')
driver.get("http://" + str(address) + "/#/login")

# Print the title of the page
title = driver.title
print(f"Page title: {title}")


# Count to keep track of the number of screenshots
screenshot_count = 0
comparison_count = 0

# Define a dictionary to map commands to functions
command_map = {
    "login": login,
    "logout": logout,
    "power_cycle": power_cycle,
    "check": check_element,
    "screenshot": screenshot,
    "check_button": check_button,
    "click": click,
    "power": power_toggle,
    "check_text": check_text,
    "wait": wait,
    "highlight": highlight,
    "tile_pattern": set_tile_pattern,
    "image_compare": compare_images,
    "element_compare": compare_elements,
    "send_serial": send_serial_command
}


current_page = "login"

PAGE_CONFIG_MAPPING = {
    "login": "login_page",
    "dashboard_button": "dashboard",
    "input_selection_button": "input_selection",
    "canvas_editor_button": "canvas_editor",
    "seam_correction_button": "seam_correction",
    "color_settings_button": "color_settings",
    "system_settings_button": "system_settings",
    "status_button": "status",
    "multiwall-control_button": "multiwall_control"
}


def set_current_page(page):
    global current_page
    current_page = page

def get_config_page_name(page):
    return PAGE_CONFIG_MAPPING.get(page, page)

def get_current_page():
    return current_page


with open('configuration.hjson', 'r') as file:
    config = hjson.load(file)


def process_command(command):
    global screenshot_count
    global comparison_count
    # Split the command into parts
    parts = command.split()
    action = parts[0]
    args = parts[1:]

    # Get the corresponding function from the dictionary
    if action in command_map:
        print(f"Processing command: {action}")

        func = command_map[action]

        try:
            if action == "login":
                username = args[0]
                password = args[1]
                func(driver, username, password)
                set_current_page("dashboard")

            elif action == "check":
                element = args[0]

                # Convert current page name to config page name
                config_page_name = get_config_page_name(current_page)
                print(f"Checking element '{element}' on page '{config_page_name}'")
                func(driver, str(element), config_page_name, config)

            elif action == "click":
                element = args[0]

                # Convert current page name to config page name
                config_page_name = get_config_page_name(current_page)
                func(driver, str(element), config_page_name, config)

                if element in PAGE_CONFIG_MAPPING:
                    new_page = PAGE_CONFIG_MAPPING[element]
                    set_current_page(new_page)
                    print(f"Navigated to {get_current_page()}")

            elif action == "screenshot":
                screenshot_count += 1
                func(driver, screenshot_count)

            elif action == "logout":
                func(driver)
                set_current_page("login")

            elif action == "power":
                func(driver)

            elif action == "power_cycle":
                proceed = args[0]
                func(driver, proceed)

            elif action == "check_button":
                button_name = args[0]
                config_page_name = get_config_page_name(current_page)
                func(driver, button_name, config_page_name, config)

            elif action == "check_text":
                element_name = args[0]
                config_page_name = get_config_page_name(current_page)
                func(driver, element_name, config_page_name, config)

            elif action == "highlight":
                element = args[0]
                config_page_name = get_config_page_name(current_page)
                print(f"Highlighting element '{element}' on page '{config_page_name}'")
                func(driver, element, config_page_name, config)

            elif action == "wait":
                duration = args[0]
                unit = args[1]
                func(int(duration), unit)

            elif action == "tile_pattern":
                option = args[0]
                func(driver, option)

            elif action == "image_compare":
                reference = args[0]
                comparison_count += 1
                func(driver, reference, comparison_count)

            elif action == "element_compare":
                element = args[0]
                reference = args[1]
                comparison_count += 1
                func(driver, element, reference, comparison_count)

            elif action == "send_serial":
                command = args[0]
                func(command, address)

            return "passed"

        except Exception as e:
            print(f"Error processing command '{command}': {e}")
            return "failed"

    else:
        print(f"Unknown command: {action}")
        return "failed"


# Get the test name and number
test_num = os.getenv('TEST_NUM')
test_name = os.getenv('TEST_NAME')

excel_filename = "combined_results.xlsx"


# Handle all the data from the different tests into one excel file
def save_excel_results(data, worksheet_name):
    try:
        # Load existing workbook or create new one
        if Path(excel_filename).exists():
            workbook = load_workbook(excel_filename)
        else:
            workbook = Workbook()
            workbook.remove(workbook.active)  # Clear default sheet

        # Create worksheet name
        sheet_name = f"Test_{test_num}_{test_name}"

        # If sheet already exists, remove it
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])

        # Create new worksheet
        worksheet = workbook.create_sheet(title=sheet_name)

        # Add headers
        headers = ['Command', 'Status']
        worksheet.append(headers)

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        # Write command results
        for row in data[:-1]:  # Exclude the summary row
            worksheet.append(row)

        # Add summary row
        worksheet.append([])
        summary = data[-1]
        worksheet.append(summary)

        # Make summary row bold
        last_row = worksheet.max_row
        for cell in worksheet[last_row]:
            cell.font = Font(bold=True)

        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Save the workbook
        workbook.save(excel_filename)
        print(f"Added/Updated worksheet '{sheet_name}' in {excel_filename}")

    except Exception as e:
        print(f"Error saving to Excel: {str(e)}")
        raise  # Re-raise the exception for debugging


results_data = []

with open('input.txt', 'r') as file:
    content = file.readlines()
    # Keep track of the number of commands given and the amount that passed
    command_count = 0
    passed_count = 0

    # Process each line as a command
    for line in content:
        command_input = line.strip()
        status = process_command(command_input)
        results_data.append([command_input, status])

        command_count += 1
        if status == "passed":
            passed_count += 1

    pass_test = 'PASSED' if passed_count == command_count else 'FAILED'

    # Add summary row
    results_data.append(['Total Passed', passed_count, 'Total Commands', command_count, 'RESULT', pass_test])

# Save results to Excel
save_excel_results(results_data, f"Test_{test_num}_{test_name}")
print(f"Test {test_name} (#{test_num}) results added to {excel_filename}")


